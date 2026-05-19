import os
import re
import csv
import time
import smtplib
import urllib.request
from datetime import date
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# ── Configuration ─────────────────────────────────────────────────────────────

GMAIL_SENDER       = os.environ["GMAIL_SENDER"]
GMAIL_APP_PASSWORD = os.environ["GMAIL_APP_PASSWORD"]
RECIPIENT_EMAIL    = os.environ["RECIPIENT_EMAIL"]
ARTISTS_SHEET_URL  = os.environ["ARTISTS_SHEET_URL"]

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
}

CONSIGNER_MAP = {
    "lougher":   "lougher",
    "clifton":   "clifton",
    "bents":     "bents",
    "fils":      "fils",
    "fontaine":  "fontaine",
    "poligrafa": "poligrafa",
}

PRICE_COLUMN = {
    "lougher":   "net",
    "clifton":   "net",
    "bents":     "net",
    "fils":      "retail",
    "poligrafa": "retail",
    "fontaine":  None,
}

ACTIVE_GALLERIES = {"lougher", "clifton", "bents", "fils", "fontaine", "poligrafa"}

# ── Helpers ───────────────────────────────────────────────────────────────────

def extract_url(cell_value):
    if not cell_value:
        return None
    text = str(cell_value)
    match = re.search(r'https?://\S+', text)
    return match.group(0).strip() if match else None


def parse_price(raw):
    if raw is None:
        return None
    text = str(raw).strip()
    if re.search(r'\d+\.\d{3},\d{2}', text):
        text = text.replace('.', '').replace(',', '.')
    elif re.search(r'\d+,\d{2}', text):
        text = text.replace(',', '.')
    text = re.sub(r'[^\d.]', '', text)
    try:
        return float(text) if text else None
    except ValueError:
        return None


def fetch_page(url):
    try:
        r = requests.get(url, headers=HEADERS, timeout=20, allow_redirects=True)
        soup = BeautifulSoup(r.text, "lxml")
        return r.status_code, soup, r.text
    except Exception:
        return None, None, None


def fetch_soup(url):
    status, soup, _ = fetch_page(url)
    if status == 200:
        return soup
    return None


def url_handle(url):
    if not url:
        return None
    path = url.lower().split("?")[0].rstrip("/")
    return path.split("/")[-1]


def load_tracked_artists():
    artists = set()
    try:
        with urllib.request.urlopen(ARTISTS_SHEET_URL, timeout=15) as resp:
            lines = resp.read().decode("utf-8").splitlines()
        reader = csv.DictReader(lines)
        for row in reader:
            name = row.get("Artist", "").strip()
            if name:
                artists.add(name.lower())
    except Exception as e:
        print(f"DEBUG: failed to load artists sheet: {e}")
    print(f"DEBUG: loaded {len(artists)} tracked artists")
    return artists


def load_consignments():
    path = os.path.join("data", next(
        f for f in os.listdir("data") if f.endswith(".xlsx")
    ))
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in ws[2]]

    def col(row, name):
        try:
            return row[headers.index(name)].value
        except (ValueError, IndexError):
            return None

    items = []
    for row in ws.iter_rows(min_row=3):
        consigner_raw = col(row, "Consigner") or ""
        gallery_key = None
        for fragment, key in CONSIGNER_MAP.items():
            if fragment in consigner_raw.lower():
                gallery_key = key
                break
        if not gallery_key:
            continue

        source_url = extract_url(col(row, "Internal Comment"))
        admin_link = col(row, "Admin Link")
        retail_raw = col(row, "Retail")
        net_raw    = col(row, "Net Price")
        price_col  = PRICE_COLUMN[gallery_key]
        our_price  = parse_price(net_raw if price_col == "net" else retail_raw)

        items.append({
            "id":         col(row, "Inventory Id"),
            "artist":     col(row, "Artist") or "",
            "title":      col(row, "Title") or "",
            "consigner":  consigner_raw,
            "gallery":    gallery_key,
            "source_url": source_url,
            "admin_link": admin_link,
            "our_price":  our_price,
        })
    return items


# ── Scrapers ──────────────────────────────────────────────────────────────────

def scrape_lougher(url):
    try:
        r = requests.get(url, headers=HEADERS, timeout=20)
        if r.status_code == 404:
            return False, None
        if r.status_code != 200:
            return None, None
        soup = BeautifulSoup(r.text, "lxml")
        for script in soup.find_all("script"):
            if not script.string:
                continue
            if '"price"' in script.string:
                m = re.search(r'"price"\s*:\s*(\d+)', script.string)
                if m:
                    price = int(m.group(1)) / 100
                    if price > 0:
                        return True, price
        return True, None
    except Exception:
        return None, None


def scrape_clifton(url):
    status, soup, _ = fetch_page(url)
    if status is None:
        return None, None
    if status == 404:
        return False, None
    if soup is None:
        return None, None
    page_text = soup.get_text()
    if re.search(r'\bSold\b', page_text):
        return False, None
    # Find price in a dedicated price element
    price_tag = soup.find(class_=re.compile(r'price', re.I))
    if price_tag:
        price_text = price_tag.get_text(strip=True)
        m = re.search(r'£([\d,]+\.?\d*)', price_text)
        if m:
            return True, parse_price(m.group(1))
    # Fallback: find first £X,XXX.XX in page
    m = re.search(r'£([\d,]{4,}\.?\d*)', page_text)
    if m:
        return True, parse_price(m.group(1))
    return True, None


def scrape_bents(url):
    status, soup, _ = fetch_page(url)
    if status is None:
        return None, None
    if status == 404:
        return False, None
    if soup is None:
        return None, None
    page_text = soup.get_text()
    if re.search(r'\bSold\b', page_text):
        return False, None
    m = re.search(r'£([\d,]{4,}\.?\d*)', page_text)
    if m:
        return True, parse_price(m.group(1))
    return True, None


def scrape_fils(url):
    status, soup, _ = fetch_page(url)
    if status is None:
        return None, None
    if status == 404:
        return False, None
    if soup is None:
        return None, None
    if not soup.find(string=re.compile(r'In den Warenkorb|inkl.*MwSt', re.I)):
        return None, None
    price_match = re.search(r'([\d]+(?:\.\d{3})*,\d{2})\s*€', soup.get_text())
    if price_match:
        return True, parse_price(price_match.group(1))
    return True, None


def scrape_fontaine(url):
    status, soup, _ = fetch_page(url)
    if status is None:
        return None, None
    if status == 404:
        return False, None
    if soup is None:
        return None, None
    return True, None


def scrape_poligrafa(url, title):
    status, soup, raw_text = fetch_page(url)
    if status is None:
        return None, None
    if status == 404:
        return False, None
    if not raw_text:
        return None, None
    if title and str(title).lower() in raw_text.lower():
        return True, None  # Skip price — too unreliable on artist pages
    return False, None


SCRAPERS = {
    "lougher":   scrape_lougher,
    "clifton":   scrape_clifton,
    "bents":     scrape_bents,
    "fils":      scrape_fils,
    "fontaine":  scrape_fontaine,
    "poligrafa": scrape_poligrafa,
}


def scrape(item):
    url = item["source_url"]
    if not url:
        return None, None
    gallery = item["gallery"]
    scraper = SCRAPERS[gallery]
    if gallery == "poligrafa":
        return scraper(url, item["title"])
    return scraper(url)


# ── New arrivals ──────────────────────────────────────────────────────────────

def check_new_arrivals(tracked_artists, consignments):
    # Build sets of what we already have per gallery
    already = {g: {"handles": set(), "titles": set()} for g in ACTIVE_GALLERIES}
    for item in consignments:
        g = item["gallery"]
        if item["source_url"]:
            already[g]["handles"].add(url_handle(item["source_url"]))
        if item["title"]:
            already[g]["titles"].add(item["title"].lower().strip())

    arrivals = []
    arrivals += _lougher_arrivals(tracked_artists, already["lougher"])
    arrivals += _clifton_arrivals(tracked_artists, already["clifton"])
    arrivals += _bents_arrivals(tracked_artists, already["bents"])
    arrivals += _fils_arrivals(tracked_artists, already["fils"])
    arrivals += _fontaine_arrivals(tracked_artists, already["fontaine"])
    arrivals += _poligrafa_arrivals(tracked_artists, already["poligrafa"])
    return arrivals


def _lougher_arrivals(tracked_artists, already):
    found = []
    seen_handles = set()
    base_url = "https://www.loughercontemporary.com/collections/all?filter.v.availability=1&filter.p.vendor=Lougher+Contemporary&sort_by=created-descending"
    page = 1
    while True:
        url = base_url if page == 1 else f"{base_url}&page={page}"
        soup = fetch_soup(url)
        if not soup:
            print(f"DEBUG: could not fetch Lougher page {page}")
            break
        links = soup.select("a[href*='/products/']")
        if not links:
            break
        new_on_page = 0
        for a in links:
            href = a.get("href", "")
            if not href:
                continue
            full_url = ("https://www.loughercontemporary.com" + href
                        if href.startswith("/") else href)
            handle = url_handle(full_url)
            if not handle or handle in already["handles"] or handle in seen_handles:
                continue
            text = a.get_text(strip=True).lower()
            if not text:
                continue
            for artist in tracked_artists:
                if artist in text:
                    seen_handles.add(handle)
                    new_on_page += 1
                    found.append({
                        "artist":  artist.title(),
                        "title":   a.get_text(strip=True),
                        "gallery": "Lougher",
                        "url":     full_url,
                    })
                    break
        next_page = soup.select_one("a[href*='page=']")
        if not next_page:
            break
        page += 1
        if page > 10:
            break
    print(f"DEBUG Lougher new arrivals: {len(found)}")
    return found


def _clifton_arrivals(tracked_artists, already):
    found = []
    # Get artist list from Clifton
    soup = fetch_soup("https://www.cliftongallery.com/artists")
    if not soup:
        print("DEBUG: could not fetch Clifton artists page")
        return found
    # Find all artist page links
    artist_links = {}
    for a in soup.select("a[href]"):
        href = a.get("href", "")
        text = a.get_text(strip=True).lower()
        if not text or not href:
            continue
        for artist in tracked_artists:
            if artist == text or artist in text:
                full_url = href if href.startswith("http") else "https://www.cliftongallery.com" + href
                artist_links[artist] = full_url
                break

    for artist, artist_url in artist_links.items():
        soup = fetch_soup(artist_url)
        if not soup:
            continue
        for a in soup.select("a[href*='/product-page/']"):
            href = a.get("href", "")
            if not href:
                continue
            full_url = href if href.startswith("http") else "https://www.cliftongallery.com" + href
            handle = url_handle(full_url)
            if handle in already["handles"]:
                continue
            title = a.get_text(strip=True)
            if not title or title.lower() in already["titles"]:
                continue
            # Check page is not sold
            status, page_soup, _ = fetch_page(full_url)
            if status == 404:
                continue
            if page_soup and re.search(r'\bSold\b', page_soup.get_text()):
                continue
            already["handles"].add(handle)
            found.append({
                "artist":  artist.title(),
                "title":   title,
                "gallery": "Clifton",
                "url":     full_url,
            })
        time.sleep(0.5)
    print(f"DEBUG Clifton new arrivals: {len(found)}")
    return found


def _bents_arrivals(tracked_artists, already):
    found = []
    for artist in tracked_artists:
        # Build Bents artist URL: firstname-lastname
        parts = artist.strip().split()
        if len(parts) < 2:
            slug = parts[0]
        else:
            slug = "-".join(parts)
        artist_url = f"https://bentscontemporary.co.uk/artist/{slug}/"
        soup = fetch_soup(artist_url)
        if not soup:
            continue
        for a in soup.select("a[href*='/artwork/']"):
            href = a.get("href", "")
            if not href:
                continue
            full_url = href if href.startswith("http") else "https://bentscontemporary.co.uk" + href
            handle = url_handle(full_url)
            if handle in already["handles"]:
                continue
            title = a.get_text(strip=True)
            if not title or title.lower() in already["titles"]:
                continue
            # Check not sold
            status, page_soup, _ = fetch_page(full_url)
            if status == 404:
                continue
            if page_soup and re.search(r'\bSold\b', page_soup.get_text()):
                continue
            already["handles"].add(handle)
            found.append({
                "artist":  artist.title(),
                "title":   title,
                "gallery": "Bents",
                "url":     full_url,
            })
        time.sleep(0.5)
    print(f"DEBUG Bents new arrivals: {len(found)}")
    return found


def _fils_arrivals(tracked_artists, already):
    found = []
    for artist in tracked_artists:
        # Build Fils artist URL: lastname-firstname
        parts = artist.strip().split()
        if len(parts) < 2:
            slug = parts[0]
        else:
            slug = f"{parts[-1]}-{'-'.join(parts[:-1])}"
        artist_url = f"https://www.fils-fine-arts.de/{slug}/"
        status, soup, _ = fetch_page(artist_url)
        if status != 200 or not soup:
            continue
        for a in soup.select("a[href*='kunst-kaufen']"):
            href = a.get("href", "")
            if not href:
                continue
            full_url = href if href.startswith("http") else "https://www.fils-fine-arts.de" + href
            handle = url_handle(full_url)
            if handle in already["handles"]:
                continue
            title = a.get_text(strip=True)
            if not title or title.lower() in already["titles"]:
                continue
            already["handles"].add(handle)
            found.append({
                "artist":  artist.title(),
                "title":   title,
                "gallery": "Fils",
                "url":     full_url,
            })
        time.sleep(0.5)
    print(f"DEBUG Fils new arrivals: {len(found)}")
    return found


def _fontaine_arrivals(tracked_artists, already):
    found = []
    # Get all artists from Fontaine artists page
    soup = fetch_soup("https://www.robertfontainegallery.com/artists/")
    if not soup:
        print("DEBUG: could not fetch Fontaine artists page")
        return found

    artist_links = {}
    for a in soup.select("a[href*='/artists/']"):
        text = a.get_text(strip=True).lower()
        href = a.get("href", "")
        if not text or not href:
            continue
        for artist in tracked_artists:
            if artist in text or text in artist:
                full_url = href if href.startswith("http") else "https://www.robertfontainegallery.com" + href
                if "/works" not in full_url:
                    full_url = full_url.rstrip("/") + "/works/"
                artist_links[artist] = full_url
                break

    for artist, artist_url in artist_links.items():
        soup = fetch_soup(artist_url)
        if not soup:
            continue
        for a in soup.select("a[href*='/works/']"):
            href = a.get("href", "")
            if not href or href == artist_url:
                continue
            full_url = href if href.startswith("http") else "https://www.robertfontainegallery.com" + href
            handle = url_handle(full_url)
            if handle in already["handles"]:
                continue
            title = a.get_text(strip=True)
            if not title or title.lower() in already["titles"]:
                continue
            # Check not 404
            status, _, _ = fetch_page(full_url)
            if status == 404:
                continue
            already["handles"].add(handle)
            found.append({
                "artist":  artist.title(),
                "title":   title,
                "gallery": "Fontaine",
                "url":     full_url,
            })
        time.sleep(0.5)
    print(f"DEBUG Fontaine new arrivals: {len(found)}")
    return found


def _poligrafa_arrivals(tracked_artists, already):
    found = []
    # Get all artists from Poligrafa artists page
    soup = fetch_soup("https://poligrafa.net/en/artistas")
    if not soup:
        print("DEBUG: could not fetch Poligrafa artists page")
        return found

    artist_links = {}
    for a in soup.select("a[href*='/artistas/']"):
        text = a.get_text(strip=True).lower()
        href = a.get("href", "")
        if not text or not href:
            continue
        for artist in tracked_artists:
            if artist in text or text in artist:
                full_url = href if href.startswith("http") else "https://poligrafa.net" + href
                artist_links[artist] = full_url
                break

    for artist, artist_url in artist_links.items():
        status, soup, raw_text = fetch_page(artist_url)
        if status != 200 or not raw_text:
            continue
        # Find all artwork titles in page source
        titles_on_page = re.findall(r'"title"\s*:\s*"([^"]+)"', raw_text)
        if not titles_on_page:
            # Try alternative pattern
            titles_on_page = re.findall(r'<h\d[^>]*>([^<]{3,80})</h\d>', raw_text)
        for title in titles_on_page:
            title_clean = title.strip()
            if not title_clean:
                continue
            if title_clean.lower() in already["titles"]:
                continue
            already["titles"].add(title_clean.lower())
            found.append({
                "artist":  artist.title(),
                "title":   title_clean,
                "gallery": "Poligrafa",
                "url":     artist_url,
            })
        time.sleep(0.5)
    print(f"DEBUG Poligrafa new arrivals: {len(found)}")
    return found


# ── Email ─────────────────────────────────────────────────────────────────────

def send_email(subject, html_body):
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"]    = GMAIL_SENDER
    msg["To"]      = RECIPIENT_EMAIL
    msg.attach(MIMEText(html_body, "html"))
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(GMAIL_SENDER, GMAIL_APP_PASSWORD)
        server.sendmail(GMAIL_SENDER, RECIPIENT_EMAIL, msg.as_string())


def build_email(today, sold, mismatches, arrivals, total_checked):
    parts = []
    if sold:       parts.append(f"{len(sold)} sold/removed")
    if mismatches: parts.append(f"{len(mismatches)} price")
    if arrivals:   parts.append(f"{len(arrivals)} new")
    summary_str = ", ".join(parts) if parts else "no changes"
    subject = f"Consignment Watch — {today.strftime('%-d %b %Y')} — {summary_str}"

    def link(url, text):
        return f'<a href="{url}">{text}</a>' if url else text

    def group_and_sort(items):
        groups = {}
        for i in items:
            key = i["consigner"] if "consigner" in i else i.get("gallery", "")
            groups.setdefault(key, []).append(i)
        for key in groups:
            groups[key].sort(key=lambda x: x.get("artist", "").lower())
        return groups

    sold_groups = group_and_sort(sold)
    rows_sold = ""
    for consigner in sorted(sold_groups.keys()):
        rows_sold += f"<li><strong>{consigner}</strong><ul>"
        for i in sold_groups[consigner]:
            rows_sold += (
                f"<li>{i['artist']} &ldquo;{i['title']}&rdquo; "
                f"[{link(i['source_url'], 'gallery')} | {link(i['admin_link'], 'admin')}]</li>"
            )
        rows_sold += "</ul></li>"

    mismatch_groups = group_and_sort(mismatches)
    rows_mismatch = ""
    for consigner in sorted(mismatch_groups.keys()):
        rows_mismatch += f"<li><strong>{consigner}</strong><ul>"
        for i in mismatch_groups[consigner]:
            rows_mismatch += (
                f"<li>{i['artist']} &ldquo;{i['title']}&rdquo;: "
                f"our {i['our_price']} vs theirs {i['their_price']} "
                f"[{link(i['source_url'], 'gallery')} | {link(i['admin_link'], 'admin')}]</li>"
            )
        rows_mismatch += "</ul></li>"

    arrival_groups = {}
    for a in arrivals:
        arrival_groups.setdefault(a["gallery"], []).append(a)
    for key in arrival_groups:
        arrival_groups[key].sort(key=lambda x: x.get("artist", "").lower())
    rows_arrivals = ""
    for gallery in sorted(arrival_groups.keys()):
        rows_arrivals += f"<li><strong>{gallery}</strong><ul>"
        for a in arrival_groups[gallery]:
            rows_arrivals += f"<li>{a['artist']} &mdash; {link(a['url'], a['title'])}</li>"
        rows_arrivals += "</ul></li>"

    html = f"""
    <p>Good morning Kris and Nana.</p>
    <p>Daily consignment report &mdash; {today.strftime('%-d %B %Y')}.</p>
    <p><strong>Summary:</strong> {total_checked} consigned artworks checked.
    {len(sold)} sold/removed &bull; {len(mismatches)} price mismatches &bull;
    {len(arrivals)} new arrivals.</p>

    <h3>🔴 Sold or removed ({len(sold)})</h3>
    <ul>{rows_sold if sold else "<li>None</li>"}</ul>

    <h3>🟡 Price mismatches ({len(mismatches)})</h3>
    <ul>{rows_mismatch if mismatches else "<li>None</li>"}</ul>

    <h3>🟢 New arrivals ({len(arrivals)})</h3>
    <ul>{rows_arrivals if arrivals else "<li>None</li>"}</ul>
    """
    return subject, html


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    today = date.today()
    consignments = load_consignments()
    tracked_artists = load_tracked_artists()

    sold       = []
    mismatches = []

    for item in consignments:
        if item["gallery"] not in ACTIVE_GALLERIES:
            continue
        if not item["source_url"]:
            continue
        found, their_price = scrape(item)
        if found is None:
            continue
        if not found:
            sold.append(item)
        elif item["our_price"] and their_price:
            diff = abs(item["our_price"] - their_price)
            pct  = diff / item["our_price"]
            if pct > 0.02:
                item["their_price"] = their_price
                mismatches.append(item)

    arrivals = check_new_arrivals(tracked_artists, consignments)
    subject, html = build_email(today, sold, mismatches, arrivals, len(consignments))
    send_email(subject, html)
    print(f"Email sent: {subject}")


if __name__ == "__main__":
    main()
